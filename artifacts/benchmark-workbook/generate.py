#!/usr/bin/env python3
"""
Benchmark Workbook Generator — first SkillForge artifact skill.

Reads benchmark results from JSON files and generates a multi-sheet
Excel workbook with summary, per-format comparison, and charts.

Usage:
    pip install openpyxl
    python artifacts/benchmark-workbook/generate.py
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
RED_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def build_document_token_sheet(wb):
    """Sheet 1: Document token efficiency from benchmarks/document-tokens/results.json."""
    results_path = PROJECT_ROOT / "benchmarks" / "document-tokens" / "results.json"
    if not results_path.exists():
        print(f"  Skipping Document Tokens — {results_path} not found")
        return
    with open(results_path) as f:
        data = json.load(f)

    ws = wb.create_sheet("Document Tokens")

    # Format key mapping: (totals key prefix, display label)
    format_defs = [
        ("raw_html", "Raw HTML"),
        ("clean_html_text", "Cleaned HTML"),
        ("raw_pdf", "Raw PDF (file)"),
        ("raw_pdf_text", "Raw PDF (text)"),
        ("raw_md", "Raw Markdown"),
        ("aif_json", "AIF JSON IR"),
        ("aif_html", "AIF HTML"),
        ("aif_md", "AIF Markdown (RT)"),
        ("aif_lml", "AIF LML Standard"),
        ("aif_lml_compact", "AIF LML Compact"),
        ("aif_lml_conserv", "AIF LML Conserv."),
        ("aif_lml_moderate", "AIF LML Moderate"),
        ("aif_lml_aggress", "AIF LML Aggress."),
    ]

    headers = ["Format", "Total Tokens", "vs Raw HTML (%)", "Total Bytes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    totals = data.get("totals", {})
    row = 2
    for key, label in format_defs:
        tokens = totals.get(f"{key}_tokens", 0)
        bytes_val = totals.get(f"{key}_bytes", 0)
        # savings percentage key varies: "savings_X_pct" or "X_save_pct"
        save_pct = totals.get(f"savings_{key}_pct", totals.get(f"{key}_save_pct", 0))

        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=tokens)
        ws.cell(row=row, column=3, value=round(save_pct, 1))
        ws.cell(row=row, column=4, value=bytes_val)

        # Color code savings
        if save_pct > 50:
            ws.cell(row=row, column=3).fill = GREEN_FILL
        elif save_pct > 5:
            ws.cell(row=row, column=3).fill = YELLOW_FILL
        elif save_pct < -5:
            ws.cell(row=row, column=3).fill = RED_FILL

        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Bar chart — token count by format
    if row > 3:
        chart = BarChart()
        chart.title = "Token Count by Format"
        chart.y_axis.title = "Tokens"
        chart.style = 10
        chart.width = 28
        chart.height = 15
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "F2")

    auto_width(ws)
    print(f"  Document Tokens: {row - 2} formats")


def build_skill_execution_sheet(wb):
    """Sheet 2: Skill execution quality from benchmarks/skill-execution/results.json."""
    results_path = PROJECT_ROOT / "benchmarks" / "skill-execution" / "results.json"
    if not results_path.exists():
        print(f"  Skipping Skill Execution — {results_path} not found")
        return
    with open(results_path) as f:
        data = json.load(f)

    ws = wb.create_sheet("Skill Execution")

    # --- Format summary table ---
    headers = ["Format", "Runs", "Avg Tokens", "Step Coverage", "Constraint Respect",
               "Output Contract", "Overall", "Min", "Max", "Stddev", "Compliance/1K Tokens"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    row = 2
    for fmt in data.get("format_summary", data.get("token_efficiency", [])):
        ws.cell(row=row, column=1, value=fmt.get("format", ""))
        ws.cell(row=row, column=2, value=fmt.get("count", 0))
        ws.cell(row=row, column=3, value=round(fmt.get("avg_tokens", 0), 1))
        ws.cell(row=row, column=4, value=round(fmt.get("avg_step_coverage", 0), 3))
        ws.cell(row=row, column=5, value=round(fmt.get("avg_constraint_respect", 0), 3))
        ws.cell(row=row, column=6, value=round(fmt.get("avg_output_contract", 0), 3))
        ws.cell(row=row, column=7, value=round(fmt.get("avg_overall", 0), 3))
        ws.cell(row=row, column=8, value=round(fmt.get("min_overall", 0), 2))
        ws.cell(row=row, column=9, value=round(fmt.get("max_overall", 0), 2))
        ws.cell(row=row, column=10, value=round(fmt.get("stddev_overall", 0), 3))
        ws.cell(row=row, column=11, value=round(fmt.get("compliance_per_1k_tokens", 0), 4))

        # Color overall
        overall = fmt.get("avg_overall", 0)
        if overall >= 0.84:
            ws.cell(row=row, column=7).fill = GREEN_FILL
        elif overall < 0.80:
            ws.cell(row=row, column=7).fill = RED_FILL
        else:
            ws.cell(row=row, column=7).fill = YELLOW_FILL

        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    summary_end = row

    # --- Difficulty breakdown ---
    difficulty = data.get("difficulty_breakdown", {})
    if difficulty:
        row += 2
        ws.cell(row=row, column=1, value="Difficulty Breakdown").font = Font(bold=True, size=12)
        row += 1
        diff_headers = ["Difficulty", "Format", "Avg Overall", "Avg Step Coverage", "Avg Constraint Respect", "Count"]
        for col, h in enumerate(diff_headers, 1):
            ws.cell(row=row, column=col, value=h)
        style_header_row(ws, row, len(diff_headers))
        row += 1

        for level in ["easy", "medium", "hard"]:
            for fmt in difficulty.get(level, []):
                ws.cell(row=row, column=1, value=level.capitalize())
                ws.cell(row=row, column=2, value=fmt.get("format", ""))
                ws.cell(row=row, column=3, value=round(fmt.get("avg_overall", 0), 3))
                ws.cell(row=row, column=4, value=round(fmt.get("avg_step_coverage", 0), 3))
                ws.cell(row=row, column=5, value=round(fmt.get("avg_constraint_respect", 0), 3))
                ws.cell(row=row, column=6, value=fmt.get("count", 0))
                for c in range(1, len(diff_headers) + 1):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1

    # Bar chart — overall score by format
    if summary_end > 3:
        chart = BarChart()
        chart.title = "Avg Overall Score by Format"
        chart.y_axis.title = "Score"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
        chart.style = 10
        chart.width = 20
        chart.height = 12
        data_ref = Reference(ws, min_col=7, min_row=1, max_row=summary_end - 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=summary_end - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "M2")

    auto_width(ws)
    print(f"  Skill Execution: {summary_end - 2} formats")


def build_roundtrip_sheet(wb):
    """Sheet 3: Roundtrip fidelity from benchmarks/roundtrip/results.json."""
    results_path = PROJECT_ROOT / "benchmarks" / "roundtrip" / "results.json"
    if not results_path.exists():
        print(f"  Skipping Roundtrip — {results_path} not found")
        return
    with open(results_path) as f:
        data = json.load(f)

    ws = wb.create_sheet("Roundtrip Fidelity")

    headers = ["File", "Format", "Blocks Orig", "Blocks RT", "Block Ratio",
               "Block Types", "Semantic Types", "Metadata", "Inline", "Overall"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    results_list = data if isinstance(data, list) else data.get("results", [])
    row = 2
    for r in results_list:
        ws.cell(row=row, column=1, value=r.get("file", ""))
        ws.cell(row=row, column=2, value=r.get("format", ""))
        ws.cell(row=row, column=3, value=r.get("block_count_original", 0))
        ws.cell(row=row, column=4, value=r.get("block_count_roundtripped", 0))
        ws.cell(row=row, column=5, value=r.get("block_count_ratio", 0))
        ws.cell(row=row, column=6, value=r.get("block_type_preservation", 0))
        ws.cell(row=row, column=7, value=r.get("semantic_type_preservation", 0))
        ws.cell(row=row, column=8, value=r.get("metadata_preservation", 0))
        ws.cell(row=row, column=9, value=r.get("inline_fidelity", 0))
        ws.cell(row=row, column=10, value=r.get("overall_fidelity", 0))

        # Color overall fidelity
        fidelity = r.get("overall_fidelity", 0)
        if fidelity >= 0.9:
            ws.cell(row=row, column=10).fill = GREEN_FILL
        elif fidelity >= 0.5:
            ws.cell(row=row, column=10).fill = YELLOW_FILL
        else:
            ws.cell(row=row, column=10).fill = RED_FILL

        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    auto_width(ws)
    print(f"  Roundtrip Fidelity: {row - 2} entries")


def build_summary_sheet(wb):
    """Sheet 0: Executive summary with key findings from all benchmarks."""
    ws = wb.active
    ws.title = "Summary"

    ws.cell(row=1, column=1, value="SkillForge Benchmark Summary")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="16213E")
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value="Generated from benchmarks/*/results.json")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

    findings = [
        ("Document cleaning", "Cleaned HTML 544K tokens vs Raw HTML 5.5M (90% reduction)"),
        ("AIF LML Aggressive", "82.2% fewer tokens than raw HTML with full semantic types"),
        ("AIF vs Raw Markdown", "LML Aggressive 981K vs Raw Markdown 1.26M (22% fewer tokens)"),
        ("Skill execution: LML", "0.844 overall — highest across all formats"),
        ("Skill execution: Markdown", "0.801 overall — 4pp below LML Aggressive"),
        ("Constraint resistance", "+11pp on hard scenarios: LML 0.79 vs Markdown 0.67"),
        ("Token efficiency", "LML: 0.97 compliance/1K tokens vs JSON IR: 0.21"),
        ("Roundtrip: JSON", "1.00 — perfect lossless roundtrip"),
        ("Roundtrip: Markdown", "~0.31 — loses semantic and block types"),
        ("Roundtrip: HTML", "0.00 — generic mode loses all structure"),
    ]

    ws.cell(row=4, column=1, value="Key Finding")
    ws.cell(row=4, column=2, value="Result")
    style_header_row(ws, 4, 2)

    for i, (finding, result) in enumerate(findings, 5):
        ws.cell(row=i, column=1, value=finding).border = THIN_BORDER
        ws.cell(row=i, column=2, value=result).border = THIN_BORDER

    # Add sheet index
    idx_row = 5 + len(findings) + 2
    ws.cell(row=idx_row, column=1, value="Sheets in This Workbook")
    ws.cell(row=idx_row, column=1).font = Font(bold=True, size=12)
    sheets = [
        ("Document Tokens", "Token counts across 13 formats for 10 Wikipedia articles"),
        ("Skill Execution", "LLM compliance scores across 4 formats, 19 scenarios"),
        ("Roundtrip Fidelity", "Format roundtrip preservation scores per file"),
    ]
    idx_row += 1
    ws.cell(row=idx_row, column=1, value="Sheet")
    ws.cell(row=idx_row, column=2, value="Description")
    style_header_row(ws, idx_row, 2)
    for name, desc in sheets:
        idx_row += 1
        ws.cell(row=idx_row, column=1, value=name).border = THIN_BORDER
        ws.cell(row=idx_row, column=2, value=desc).border = THIN_BORDER

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 65


def main():
    wb = Workbook()

    print("Building workbook...")
    build_summary_sheet(wb)
    build_document_token_sheet(wb)
    build_skill_execution_sheet(wb)
    build_roundtrip_sheet(wb)

    output_dir = Path(__file__).parent
    output_path = output_dir / "skillforge_benchmarks.xlsx"
    wb.save(str(output_path))
    print(f"\nWorkbook saved to {output_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
